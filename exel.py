import openpyxl

filename = "data-2.xlsx"

# load the workbook
workbook = openpyxl.load_workbook(filename)

# select the worksheet
worksheet = workbook.active

# read data from cells
# cell_value = worksheet["A1"].value
# print(cell_value)

# read data from rows
# for row in worksheet.iter_rows(min_row=2, max_col=3):
#     values = [cell.value for cell in row]
#     print(values)

# read data from columns
for column in worksheet.iter_cols(min_col=1, max_row=3):
    values = [cell.value for cell in column]
    print(values)

# read data from a range of cells
# cell_range = worksheet["A1:C3"]
# for row in cell_range:
#     values = [cell.value for cell in row]
#     print(values)